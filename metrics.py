import os
import re
import csv
import json
import math
import requests
import pandas as pd
from datetime import datetime, timezone
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
# CONFIG (set via environment variables)
# ============================================================
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "").strip()
GITHUB_OWNER = os.getenv("GITHUB_OWNER", "").strip()
GITHUB_REPO = os.getenv("GITHUB_REPO", "").strip()

MAIN_BRANCH = os.getenv("MAIN_BRANCH", "main").strip()
DAYS_LOOKBACK = int(os.getenv("DAYS_LOOKBACK", "30"))

OUT_DIR = os.getenv("OUT_DIR", "out").strip()
os.makedirs(OUT_DIR, exist_ok=True)

# Heuristics to identify production deployment workflows/runs
# Added 'ci' so that CI workflow runs count as "prod" for DORA metrics
PROD_NAME_REGEX = re.compile(r"(deploy|prod|production|release|ci)", re.IGNORECASE)
CI_NAME_REGEX = re.compile(r"(ci|build|test|lint|quality|gate)", re.IGNORECASE)


# ============================================================
# Helpers
# ============================================================
def iso_to_dt(s: str) -> datetime:
    return datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone(timezone.utc)

def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def within_lookback(dt: datetime, days: int) -> bool:
    return (datetime.now(timezone.utc) - dt).days <= days

def gh_get(url: str, params=None) -> dict:
    if not GITHUB_TOKEN:
        raise RuntimeError("Missing GITHUB_TOKEN env var.")
    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    r = requests.get(url, headers=headers, params=params, timeout=30)
    if r.status_code != 200:
        raise RuntimeError(f"GitHub API error {r.status_code}: {r.text[:250]}")
    return r.json()

def paginate(url: str, params=None, max_pages: int = 5, item_key: str = None) -> list:
    """Generic GitHub pagination helper. If item_key is provided, collects data[item_key]."""
    all_items = []
    for page in range(1, max_pages + 1):
        p = dict(params or {})
        p["per_page"] = 100
        p["page"] = page
        data = gh_get(url, params=p)
        items = data.get(item_key) if item_key else data
        if not items:
            break
        all_items.extend(items)
    return all_items

def write_csv(path: str, rows: list, header=None):
    if not rows:
        # Write empty file with header if available
        with open(path, "w", newline="", encoding="utf-8") as f:
            if header:
                csv.DictWriter(f, fieldnames=header).writeheader()
        return
    if header is None:
        header = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=header)
        w.writeheader()
        w.writerows(rows)

def is_prod_run(run: dict) -> bool:
    name = (run.get("name") or "") + " " + (run.get("display_title") or "")
    return bool(PROD_NAME_REGEX.search(name))

def classify_run(run: dict) -> str:
    """Classify run into CI / Deploy / Infra."""
    conclusion = (run.get("conclusion") or "").lower()
    name = (run.get("name") or "") + " " + (run.get("display_title") or "")

    if conclusion == "startup_failure":
        return "infra"

    if PROD_NAME_REGEX.search(name):
        return "deploy"

    if CI_NAME_REGEX.search(name):
        return "ci"

    return "ci"


# ============================================================
# Fetch telemetry
# ============================================================
def fetch_workflow_runs() -> list:
    url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/actions/runs"
    return paginate(url, params={}, max_pages=5, item_key="workflow_runs")

def fetch_pull_requests() -> list:
    url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/pulls"
    # include closed, we’ll filter by lookback
    return paginate(url, params={"state": "all"}, max_pages=5, item_key=None)

def fetch_commits() -> list:
    url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/commits"
    return paginate(url, params={}, max_pages=5, item_key=None)


# ============================================================
# Metrics: DORA (prod filtered) + pipeline health + flow
# ============================================================
def dora_metrics(runs: list) -> dict:
    """
    DORA from GitHub Actions runs.
    Filters: main branch + lookback + prod-only heuristic.
    Lead Time: pipeline-time proxy (created_at -> updated_at for successful prod runs).
    MTTR: failure->next success (best-effort) for prod runs.
    """
    prod_runs = []
    for r in runs:
        created = iso_to_dt(r["created_at"])
        if r.get("head_branch") != MAIN_BRANCH:
            continue
        if not within_lookback(created, DAYS_LOOKBACK):
            continue
        if not is_prod_run(r):
            continue
        prod_runs.append(r)

    if not prod_runs:
        return {
            "deployment_frequency_per_day": 0.0,
            "lead_time_hours_proxy": 0.0,
            "change_failure_rate_pct": 0.0,
            "mttr_hours": 0.0,
            "prod_runs": 0,
            "note": "No prod runs detected. Rename deploy workflow to include deploy/prod/production/release OR adjust PROD_NAME_REGEX."
        }

    # deployment frequency
    created_times = sorted([iso_to_dt(r["created_at"]) for r in prod_runs])
    window_days = max(1, (created_times[-1] - created_times[0]).days + 1)

    success = [r for r in prod_runs if (r.get("conclusion") or "").lower() == "success"]
    failure = [r for r in prod_runs if (r.get("conclusion") or "").lower() in ("failure", "cancelled", "timed_out", "startup_failure")]

    dep_freq = len(success) / window_days
    cfr = (len(failure) / max(1, len(prod_runs))) * 100.0

    # lead time proxy = average pipeline duration
    lead_list = []
    for r in success:
        start = iso_to_dt(r["created_at"])
        end = iso_to_dt(r["updated_at"])
        lead_list.append((end - start).total_seconds() / 3600.0)
    lead = sum(lead_list) / len(lead_list) if lead_list else 0.0

    # mttr best-effort: prod failure -> next prod success
    succ_times = sorted([iso_to_dt(r["created_at"]) for r in success])
    fail_times = sorted([iso_to_dt(r["created_at"]) for r in failure])

    mttr_list = []
    for ft in fail_times:
        nxt = next((st for st in succ_times if st > ft), None)
        if nxt:
            mttr_list.append((nxt - ft).total_seconds() / 3600.0)
    mttr = sum(mttr_list) / len(mttr_list) if mttr_list else 0.0

    return {
        "deployment_frequency_per_day": round(dep_freq, 3),
        "lead_time_hours_proxy": round(lead, 2),
        "change_failure_rate_pct": round(cfr, 2),
        "mttr_hours": round(mttr, 2),
        "prod_runs": len(prod_runs),
        "note": "Lead Time is proxy (pipeline duration). For true lead time: PR merge -> next prod deploy."
    }

def dora_daily_timeseries(runs: list) -> list:
    """Daily prod success/failure counts (main + lookback + prod-only)."""
    series = defaultdict(lambda: {"success": 0, "failure": 0})
    for r in runs:
        created = iso_to_dt(r["created_at"])
        if r.get("head_branch") != MAIN_BRANCH:
            continue
        if not within_lookback(created, DAYS_LOOKBACK):
            continue
        if not is_prod_run(r):
            continue

        day = created.date().isoformat()
        concl = (r.get("conclusion") or "").lower()
        if concl == "success":
            series[day]["success"] += 1
        elif concl in ("failure", "cancelled", "timed_out", "startup_failure"):
            series[day]["failure"] += 1

    out = []
    for day in sorted(series.keys()):
        total = series[day]["success"] + series[day]["failure"]
        fr = (series[day]["failure"] / total * 100.0) if total else 0.0
        out.append({
            "day": day,
            "prod_success": series[day]["success"],
            "prod_failure": series[day]["failure"],
            "failure_rate_pct": round(fr, 2),
            "deployments": series[day]["success"]
        })
    return out

def pipeline_health(runs: list) -> list:
    """Breakdown of CI vs Deploy vs Infra conclusions (main + lookback)."""
    breakdown = defaultdict(int)
    for r in runs:
        created = iso_to_dt(r["created_at"])
        if r.get("head_branch") != MAIN_BRANCH:
            continue
        if not within_lookback(created, DAYS_LOOKBACK):
            continue
        bucket = classify_run(r)
        conclusion = (r.get("conclusion") or "").lower() or "unknown"
        breakdown[(bucket, conclusion)] += 1

    rows = []
    for (bucket, conclusion), count in sorted(breakdown.items()):
        rows.append({"bucket": bucket, "conclusion": conclusion, "count": count})
    return rows

def flow_metrics(prs: list) -> tuple[list, dict]:
    """
    Flow metrics:
    - WIP: count open PRs (within lookback window)
    - Cycle Time: PR created -> merged (hours)
    NOTE: true "PR review time" needs PR reviews endpoint; can be added later.
    """
    rows = []
    wip = 0

    for pr in prs:
        created = iso_to_dt(pr["created_at"])
        if not within_lookback(created, DAYS_LOOKBACK):
            continue

        state = pr.get("state", "")
        if state == "open":
            wip += 1

        merged_at = pr.get("merged_at")
        merged_dt = iso_to_dt(merged_at) if merged_at else None
        cycle_hours = (merged_dt - created).total_seconds() / 3600.0 if merged_dt else None

        rows.append({
            "pr_number": pr["number"],
            "title": pr.get("title", ""),
            "state": state,
            "created_at": created.isoformat(),
            "merged_at": merged_dt.isoformat() if merged_dt else "",
            "cycle_time_hours": round(cycle_hours, 2) if cycle_hours is not None else ""
        })

    return rows, {"wip_open_prs": wip}

def read_trivy_summary(trivy_path: str) -> dict | None:
    """Optional: parse Trivy JSON if present as out/trivy.json."""
    if not os.path.exists(trivy_path):
        return None
    with open(trivy_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    sev = defaultdict(int)
    for r in data.get("Results", []):
        for v in (r.get("Vulnerabilities") or []):
            sev[(v.get("Severity") or "UNKNOWN").upper()] += 1

    total = sum(sev.values())
    return {
        "trivy_total": total,
        "trivy_critical": sev.get("CRITICAL", 0),
        "trivy_high": sev.get("HIGH", 0),
        "trivy_medium": sev.get("MEDIUM", 0),
        "trivy_low": sev.get("LOW", 0),
        "trivy_unknown": sev.get("UNKNOWN", 0),
    }


# ============================================================
# Export: CSV + Excel Dashboard
# ============================================================
def write_dashboard_xlsx(path: str,
                         snapshot_df: pd.DataFrame,
                         dora_daily_df: pd.DataFrame,
                         flow_df: pd.DataFrame,
                         pipeline_df: pd.DataFrame):
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    for r in dataframe_to_rows(snapshot_df, index=False, header=True):
        ws.append(r)

    # DORA Daily
    ws2 = wb.create_sheet("DORA_Daily")
    for r in dataframe_to_rows(dora_daily_df, index=False, header=True):
        ws2.append(r)

    # Flow
    ws3 = wb.create_sheet("Flow_PRs")
    for r in dataframe_to_rows(flow_df, index=False, header=True):
        ws3.append(r)

    # Pipeline health
    ws4 = wb.create_sheet("Pipeline_Health")
    for r in dataframe_to_rows(pipeline_df, index=False, header=True):
        ws4.append(r)

    wb.save(path)


# ============================================================
# MAIN
# ============================================================
def main():
    if not GITHUB_OWNER or not GITHUB_REPO:
        raise RuntimeError("Set env vars: GITHUB_OWNER, GITHUB_REPO, and GITHUB_TOKEN.")

    # Fetch telemetry
    runs = fetch_workflow_runs()
    prs = fetch_pull_requests()
    _ = fetch_commits()  # available for later; not used in calculations here

    # Compute metrics
    dora = dora_metrics(runs)
    dora_daily = dora_daily_timeseries(runs)
    flow_rows, wip = flow_metrics(prs)
    pipe_rows = pipeline_health(runs)

    # Optional security summary
    trivy = read_trivy_summary(os.path.join(OUT_DIR, "trivy.json"))

    # Snapshot table
    snapshot = [
        {"Metric": "Deployment Frequency (per day)", "Value": dora["deployment_frequency_per_day"], "Timestamp": now_utc_iso()},
        {"Metric": "Lead Time (hours) [proxy]", "Value": dora["lead_time_hours_proxy"], "Timestamp": now_utc_iso()},
        {"Metric": "Change Failure Rate (%)", "Value": dora["change_failure_rate_pct"], "Timestamp": now_utc_iso()},
        {"Metric": "MTTR (hours)", "Value": dora["mttr_hours"], "Timestamp": now_utc_iso()},
        {"Metric": "WIP (Open PRs)", "Value": wip["wip_open_prs"], "Timestamp": now_utc_iso()},
        {"Metric": "Note", "Value": dora.get("note", ""), "Timestamp": now_utc_iso()},
    ]
    if trivy:
        snapshot.extend([
            {"Metric": "Trivy Total Vulns", "Value": trivy["trivy_total"], "Timestamp": now_utc_iso()},
            {"Metric": "Trivy Critical", "Value": trivy["trivy_critical"], "Timestamp": now_utc_iso()},
            {"Metric": "Trivy High", "Value": trivy["trivy_high"], "Timestamp": now_utc_iso()},
        ])

    # Export CSVs
    write_csv(os.path.join(OUT_DIR, "metrics_snapshot.csv"), snapshot, header=["Metric", "Value", "Timestamp"])
    write_csv(os.path.join(OUT_DIR, "dora_daily.csv"), dora_daily, header=["day", "prod_success", "prod_failure", "failure_rate_pct", "deployments"])
    write_csv(os.path.join(OUT_DIR, "flow_pr.csv"), flow_rows)
    write_csv(os.path.join(OUT_DIR, "pipeline_health.csv"), pipe_rows, header=["bucket", "conclusion", "count"])

    # Export Excel dashboard
    snapshot_df = pd.DataFrame(snapshot)
    dora_daily_df = pd.DataFrame(dora_daily)
    flow_df = pd.DataFrame(flow_rows)
    pipeline_df = pd.DataFrame(pipe_rows)

    dashboard_path = os.path.join(OUT_DIR, "dashboard.xlsx")
    write_dashboard_xlsx(dashboard_path, snapshot_df, dora_daily_df, flow_df, pipeline_df)

    # Print summary
    print("\n=== Metrics Generated ===")
    print(f"Repo: {GITHUB_OWNER}/{GITHUB_REPO} | Branch: {MAIN_BRANCH} | Lookback: {DAYS_LOOKBACK} days")
    print(f"Prod runs detected: {dora['prod_runs']}")
    print(f"Deployment Frequency/day: {dora['deployment_frequency_per_day']}")
    print(f"Lead Time (hours) [proxy]: {dora['lead_time_hours_proxy']}")
    print(f"Change Failure Rate (%): {dora['change_failure_rate_pct']}")
    print(f"MTTR (hours): {dora['mttr_hours']}")
    print(f"WIP (open PRs): {wip['wip_open_prs']}")
    print(f"Outputs written to: {OUT_DIR}/")
    print("- metrics_snapshot.csv")
    print("- dora_daily.csv")
    print("- flow_pr.csv")
    print("- pipeline_health.csv")
    print("- dashboard.xlsx")
    if trivy:
        print("- trivy.json (read and summarized)")

if __name__ == "__main__":
    main()
